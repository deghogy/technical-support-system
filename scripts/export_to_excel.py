#!/usr/bin/env python3
"""
Supabase Database to Excel Exporter

This script connects to Supabase, fetches all data from site_visit_requests and customer_quotas tables,
and exports them to an Excel file with multiple sheets and proper formatting.

Sheets:
    - Site Visit Requests: All site visit request data
    - Customer Quotas: Customer quota information with calculated available hours

Usage:
    python export_to_excel.py

Environment Variables:
    SUPABASE_URL - Your Supabase project URL
    SUPABASE_SERVICE_KEY - Your Supabase service role key (needed for full access)

Scheduling (Weekly):
    - Linux/Mac (cron): 0 0 * * 0 python /path/to/export_to_excel.py
    - Windows (Task Scheduler): Create weekly task
"""

import os
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

# Load environment variables from .env file
def load_env_file():
    """Load environment variables from .env file if it exists."""
    env_path = Path(__file__).parent / '.env'
    if env_path.exists():
        print(f"Loading environment from: {env_path}")
        with open(env_path, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    os.environ[key] = value
        return True
    return False

try:
    from supabase import create_client
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError as e:
    print(f"Missing required package: {e}")
    print("Please install dependencies: pip install -r requirements.txt")
    sys.exit(1)

# Load .env file before anything else
load_env_file()


def get_supabase_client():
    """Create and return Supabase client using environment variables."""
    supabase_url = os.environ.get('SUPABASE_URL')
    supabase_key = os.environ.get('SUPABASE_SERVICE_KEY')

    if not supabase_url or not supabase_key:
        print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY environment variables must be set")
        print("\nExample:")
        print("  export SUPABASE_URL=https://your-project.supabase.co")
        print("  export SUPABASE_SERVICE_KEY=your-service-key")
        sys.exit(1)

    return create_client(supabase_url, supabase_key)


def fetch_all_data(supabase) -> list:
    """Fetch all records from site_visit_requests table."""
    print("Fetching site visit requests...")

    # Fetch all records (adjust limit if you have more than 10,000)
    response = supabase.table('site_visit_requests').select('*').limit(10000).execute()

    if hasattr(response, 'data'):
        data = response.data
    else:
        data = response

    print(f"Fetched {len(data)} records")
    return data


def fetch_quotas_data(supabase) -> list:
    """Fetch all records from customer_quotas table."""
    print("Fetching customer quotas...")

    response = supabase.table('customer_quotas').select('*').limit(1000).execute()

    if hasattr(response, 'data'):
        data = response.data
    else:
        data = response

    print(f"Fetched {len(data)} quota records")
    return data


def format_data_for_excel(data: list) -> pd.DataFrame:
    """Convert raw data to a pandas DataFrame with proper column formatting."""
    if not data:
        return pd.DataFrame()

    df = pd.DataFrame(data)

    # Rename columns to be more readable
    column_mapping = {
        'id': 'Request ID',
        'requester_name': 'Requester Name',
        'requester_email': 'Requester Email',
        'site_location': 'Site Location',
        'problem_desc': 'Problem Description',
        'requested_date': 'Requested Date',
        'duration_hours': 'Planned Hours',
        'status': 'Request Status',
        'approved_at': 'Approved At',
        'scheduled_date': 'Scheduled Date',
        'actual_start_time': 'Actual Start Time',
        'actual_end_time': 'Actual End Time',
        'technician_notes': 'Technician Notes',
        'customer_notes': 'Customer Notes',
        'visit_status': 'Visit Status',
        'customer_confirmed_at': 'Customer Confirmed At',
        'rejection_reason': 'Rejection Reason',
        'created_at': 'Created At',
        'document_url': 'Document URL',
    }

    # Only rename columns that exist
    existing_columns = {k: v for k, v in column_mapping.items() if k in df.columns}
    df = df.rename(columns=existing_columns)

    # Reorder columns for better readability
    preferred_order = [
        'Request ID',
        'Created At',
        'Requester Name',
        'Requester Email',
        'Site Location',
        'Problem Description',
        'Request Status',
        'Visit Status',
        'Requested Date',
        'Planned Hours',
        'Approved At',
        'Scheduled Date',
        'Actual Start Time',
        'Actual End Time',
        'Customer Confirmed At',
        'Technician Notes',
        'Customer Notes',
        'Rejection Reason',
        'Document URL',
    ]

    # Only include columns that exist
    final_columns = [col for col in preferred_order if col in df.columns]
    df = df[final_columns]

    return df


def style_worksheet(ws, df: pd.DataFrame, header_color: str = '0077C8'):
    """Apply styling to a single worksheet."""
    # Define styles
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Set width with some padding (max 50 for readability)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add borders to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border


def style_excel_file(filepath: str, df_requests: pd.DataFrame, df_quotas: pd.DataFrame = None):
    """Apply styling to the Excel file with multiple sheets."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath)

    # Style Site Visit Requests sheet
    ws_requests = wb['Site Visit Requests']
    style_worksheet(ws_requests, df_requests, header_color='0077C8')

    # Style Customer Quotas sheet if it exists
    if df_quotas is not None and not df_quotas.empty:
        ws_quotas = wb['Customer Quotas']
        style_worksheet(ws_quotas, df_quotas, header_color='22C55E')

    wb.save(filepath)
    print(f"Styling applied to {filepath}")


def format_quotas_for_excel(data: list) -> pd.DataFrame:
    """Convert quotas data to a pandas DataFrame with proper formatting."""
    if not data:
        return pd.DataFrame()

    df = pd.DataFrame(data)

    # Rename columns to be more readable
    column_mapping = {
        'id': 'ID',
        'customer_email': 'Customer Email',
        'total_hours': 'Total Hours Quota',
        'used_hours': 'Used Hours',
        'created_at': 'Created At',
        'updated_at': 'Updated At',
    }

    # Only rename columns that exist
    existing_columns = {k: v for k, v in column_mapping.items() if k in df.columns}
    df = df.rename(columns=existing_columns)

    # Calculate available hours
    if 'Total Hours Quota' in df.columns and 'Used Hours' in df.columns:
        df['Available Hours'] = df['Total Hours Quota'] - df['Used Hours']

    # Reorder columns
    preferred_order = [
        'Customer Email',
        'Total Hours Quota',
        'Used Hours',
        'Available Hours',
        'Created At',
        'Updated At',
    ]

    # Only include columns that exist
    final_columns = [col for col in preferred_order if col in df.columns]
    df = df[final_columns]

    return df


def export_to_excel(data: list, quotas_data: list, output_dir: Optional[str] = None) -> str:
    """Export data to Excel file with multiple sheets."""
    if not data:
        print("No site visit data to export")
        return ""

    # Create output directory if it doesn't exist
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(__file__), 'exports')

    os.makedirs(output_dir, exist_ok=True)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"site_visit_requests_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # Format data
    df_requests = format_data_for_excel(data)
    df_quotas = format_quotas_for_excel(quotas_data)

    # Export to Excel with multiple sheets
    print(f"Exporting to {filepath}...")
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Sheet 1: Site Visit Requests
        df_requests.to_excel(writer, sheet_name='Site Visit Requests', index=False)

        # Sheet 2: Customer Quotas
        if not df_quotas.empty:
            df_quotas.to_excel(writer, sheet_name='Customer Quotas', index=False)

    # Apply styling to both sheets
    style_excel_file(filepath, df_requests, df_quotas)

    print(f"✓ Export complete: {filepath}")
    return filepath


def generate_summary(data: list) -> str:
    """Generate a text summary of the data."""
    if not data:
        return "No data available"

    total = len(data)
    pending = sum(1 for r in data if r.get('status') == 'pending')
    approved = sum(1 for r in data if r.get('status') == 'approved')
    rejected = sum(1 for r in data if r.get('status') == 'rejected')
    confirmed = sum(1 for r in data if r.get('visit_status') == 'confirmed')
    scheduled = sum(1 for r in data if r.get('scheduled_date') and r.get('status') == 'approved' and r.get('visit_status') != 'confirmed')

    summary = f"""
╔════════════════════════════════════════════════════════╗
║           SUPABASE EXPORT SUMMARY                      ║
╠════════════════════════════════════════════════════════╣
║ Total Records:        {total:>5}                          ║
║ Pending Requests:     {pending:>5}                          ║
║ Approved (not sched): {approved - scheduled - confirmed:>5}                          ║
║ Scheduled:            {scheduled:>5}                          ║
║ Completed:            {confirmed:>5}                          ║
║ Rejected:             {rejected:>5}                          ║
╚════════════════════════════════════════════════════════╝
"""
    return summary


def main():
    """Main execution function."""
    print("=" * 60)
    print("SUPABASE TO EXCEL EXPORTER")
    print("=" * 60)
    print()

    try:
        # Connect to Supabase
        supabase = get_supabase_client()

        # Fetch data
        data = fetch_all_data(supabase)
        quotas_data = fetch_quotas_data(supabase)

        # Print summary
        print(generate_summary(data))

        # Export to Excel
        if data:
            output_path = export_to_excel(data, quotas_data)

            # Also create a "latest" copy
            if output_path:
                output_dir = os.path.dirname(output_path)
                latest_path = os.path.join(output_dir, 'site_visit_requests_latest.xlsx')

                df_requests = format_data_for_excel(data)
                df_quotas = format_quotas_for_excel(quotas_data)

                with pd.ExcelWriter(latest_path, engine='openpyxl') as writer:
                    df_requests.to_excel(writer, sheet_name='Site Visit Requests', index=False)
                    if not df_quotas.empty:
                        df_quotas.to_excel(writer, sheet_name='Customer Quotas', index=False)

                style_excel_file(latest_path, df_requests, df_quotas)
                print(f"✓ Latest copy updated: {latest_path}")
        else:
            print("No records found in the database")

        print()
        print("Export completed successfully!")
        return 0

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
